#!/usr/bin/env python3
"""
Generate Excel Workbooks for H.A - Limpeza:
1. docs/03_HA_Precario.xlsx
2. docs/04_HA_Projecoes_Financeiras.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DOCS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'docs')

def create_precario_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "Preçário B2B"

    # Styles
    navy_fill = PatternFill(start_color="071B33", end_color="071B33", fill_type="solid")
    indigo_fill = PatternFill(start_color="464FEB", end_color="464FEB", fill_type="solid")
    light_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    white_bold = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    navy_bold_large = Font(name="Segoe UI", size=14, bold=True, color="071B33")
    navy_regular = Font(name="Segoe UI", size=10, color="071B33")
    regular_font = Font(name="Segoe UI", size=10)
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # Title Banner
    ws.merge_cells('A1:G1')
    ws['A1'] = "H.A - LIMPEZA | PROPERTY READINESS & ASSET CARE"
    ws['A1'].font = Font(name="Segoe UI", size=13, bold=True, color="FFFFFF")
    ws['A1'].fill = navy_fill
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:G2')
    ws['A2'] = "Tabela Tarifária Oficial B2B e Estrutura de Custos Operacionais (AML)"
    ws['A2'].font = Font(name="Segoe UI", size=10, italic=True, color="475569")
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Headers
    headers = [
        "Código", "Designação do Serviço", "Tipologia / Segmento", "Área Ref.",
        "Preço Base (€ sem IVA)", "Horas Médias Estimadas", "N.º Técnicos Alocados"
    ]
    
    ws.append([])
    ws.append(headers)
    ws.row_dimensions[4].height = 25
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = white_bold
        cell.fill = indigo_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data Rows
    data = [
        ["PP-T1", "Property Prep (Preparação Comercial)", "T0 / T1", "<= 65 m²", 280.0, 3.5, 2],
        ["PP-T2", "Property Prep (Preparação Comercial)", "T2", "<= 95 m²", 390.0, 4.5, 2],
        ["PP-T3", "Property Prep (Preparação Comercial)", "T3", "<= 130 m²", 490.0, 5.5, 2],
        ["PP-T4", "Property Prep (Preparação Comercial)", "T4+ / Moradia", "> 130 m² (preço/m²)", 3.80, 7.0, 3],
        ["RN-T1", "Renovation (Limpeza Pós-Obra)", "T0 / T1", "<= 65 m²", 420.0, 5.0, 2],
        ["RN-T2", "Renovation (Limpeza Pós-Obra)", "T2", "<= 95 m²", 580.0, 6.5, 2],
        ["RN-T3", "Renovation (Limpeza Pós-Obra)", "T3", "<= 130 m²", 750.0, 8.0, 2],
        ["RN-T4", "Renovation (Limpeza Pós-Obra)", "T4+ / Moradia", "> 130 m² (preço/m²)", 5.50, 9.5, 3],
        ["RO-T1", "Ready to Occupy (Entrega Chave)", "T0 / T1", "<= 65 m²", 320.0, 4.0, 2],
        ["RO-T2", "Ready to Occupy (Entrega Chave)", "T2", "<= 95 m²", 440.0, 5.0, 2],
        ["RO-T3", "Ready to Occupy (Entrega Chave)", "T3", "<= 130 m²", 560.0, 6.0, 2],
        ["RO-T4", "Ready to Occupy (Entrega Chave)", "T4+ / Moradia", "> 130 m² (preço/m²)", 4.20, 7.5, 3],
        ["PC-T1", "Property Care (Preservação Mensal)", "T0 / T1", "<= 65 m²", 120.0, 3.0, 1],
        ["PC-T2", "Property Care (Preservação Mensal)", "T2", "<= 95 m²", 160.0, 4.0, 1],
        ["PC-T3", "Property Care (Preservação Mensal)", "T3", "<= 130 m²", 200.0, 5.0, 1],
        ["PC-T4", "Property Care (Preservação Mensal)", "T4+ / Moradia", "> 130 m²", 240.0, 6.0, 1],
    ]

    for row_idx, row_data in enumerate(data, start=5):
        ws.append(row_data)
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx in [1, 3, 4, 7]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in [5, 6]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if col_idx == 5:
                    cell.number_format = '€#,##0.00'
                    cell.font = bold_font
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Column Widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    out_path = os.path.join(DOCS_DIR, "03_HA_Precario.xlsx")
    wb.save(out_path)
    print(f"Generated: {out_path}")

def create_projecoes_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "Projeções 5 Anos"

    navy_fill = PatternFill(start_color="071B33", end_color="071B33", fill_type="solid")
    indigo_fill = PatternFill(start_color="464FEB", end_color="464FEB", fill_type="solid")
    emerald_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    light_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    white_bold = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    regular_font = Font(name="Segoe UI", size=10)
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Title Banner
    ws.merge_cells('A1:F1')
    ws['A1'] = "H.A - LIMPEZA | MODELO FINANCEIRO PLURIANUAL (2026 - 2031)"
    ws['A1'].font = Font(name="Segoe UI", size=13, bold=True, color="FFFFFF")
    ws['A1'].fill = navy_fill
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    headers = [
        "Rubrica Financeira", "Ano 1", "Ano 2", "Ano 3", "Ano 4", "Ano 5"
    ]
    ws.append([])
    ws.append(headers)
    ws.row_dimensions[3].height = 24
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = white_bold
        cell.fill = indigo_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rows = [
        ["Intervenções / Mês", 30, 60, 100, 150, 220],
        ["Intervenções Totais / Ano", 360, 720, 1200, 1800, 2640],
        ["Ticket Médio (€)", 420.0, 420.0, 420.0, 420.0, 420.0],
        ["Receita Bruta Total (€)", 151200.0, 302400.0, 504000.0, 756000.0, 1108800.0],
        ["Custos Diretos - COGS (~38.5%)", 58140.0, 116280.0, 193800.0, 290700.0, 426360.0],
        ["Margem Bruta (€)", 93060.0, 186120.0, 310200.0, 465300.0, 682440.0],
        ["Margem Bruta (%)", 0.615, 0.615, 0.615, 0.615, 0.615],
        ["Custos Fixos / OPEX Anual (€)", 36000.0, 65000.0, 110000.0, 165000.0, 230000.0],
        ["EBITDA Operacional (€)", 57060.0, 121120.0, 200200.0, 300300.0, 452440.0],
        ["Margem EBITDA (%)", 0.377, 0.401, 0.397, 0.397, 0.408],
    ]

    for row_idx, r in enumerate(rows, start=4):
        ws.append(r)
        ws.row_dimensions[row_idx].height = 20
        is_highlight = row_idx in [7, 9, 12]
        for col_idx in range(1, len(r) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = bold_font if is_highlight else regular_font
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if row_idx in [4, 5]:
                    cell.number_format = '#,##0'
                elif row_idx in [6, 7, 8, 9, 11, 12]:
                    cell.number_format = '€#,##0.00'
                elif row_idx in [10, 13]:
                    cell.number_format = '0.0%'

    # Column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    out_path = os.path.join(DOCS_DIR, "04_HA_Projecoes_Financeiras.xlsx")
    wb.save(out_path)
    print(f"Generated: {out_path}")

if __name__ == "__main__":
    os.makedirs(DOCS_DIR, exist_ok=True)
    create_precario_sheet()
    create_projecoes_sheet()
    print("All spreadsheets created successfully.")
